import streamlit as st
from PIL import Image
import pandas as pd
import re
from fuzzywuzzy import fuzz
import difflib
import time
import openpyxl
import conflitto
import os

# --- APP SETUP ---
st.set_page_config(page_title="HMV Fair Quote Tool", layout="wide", page_icon="🔧")

# --- CSS for styling and animation ---
st.markdown("""
    <style>
    body {background: #fcfcfc;}
    .conclusion-animated {
        animation: popfadeIn 0.8s cubic-bezier(.18,.71,.53,.93);
        box-shadow: 0 0 18px 0 #e4f4e9, 0 12px 24px #b1cdfb41;
        background: linear-gradient(95deg,#f3fffa 60%,#eaf3fc 100%);
        border-left: 8px solid #38c976;
        border-radius:16px; padding:2em 2.3em; margin:1.6em 0 1.3em 0;
        position:relative;
    }
    .conclusion-animated.warn {border-left-color:#ffb338;background:linear-gradient(95deg,#fffbe9 60%,#fff8f1 100%);}
    .conclusion-animated.fail {border-left-color:#ed5565;background:linear-gradient(95deg,#fff0f0 60%,#fff8f1 100%);}
    @keyframes popfadeIn {0% {scale:0.95;opacity:0;} 100% {scale:1;opacity:1;}}
    .conclusion-head { font-size:1.35rem; font-weight:600; margin-bottom:7px;}
    .metric-card {
        background:#fff; border-radius:11px; padding:2em 1.4em 1.3em 1.4em; margin-top:18px;
        box-shadow:0 2px 11px 1px #ddeafc55; display:flex; flex-direction:column; align-items:center;
    }
    .metric-label { color:#7f8c8d; font-size:1.05rem;margin-bottom:0.3em;}
    .metric-value { font-size: 2.17rem; font-weight:700; color:#13202A;}
    .diff-positive { color: #e74c3c;}
    .diff-negative { color: #2ecc71;}
    .diff-neutral { color: #3498db;}
    .result-table th {background:#3498db; color:#fff; font-weight:600; padding:10px 12px;}
    .result-table td {background:#fafcff; padding:10px 12px;}
    .result-table tr {border-bottom:1px solid #dde9f7;}
    .pulse-anim {
       animation: pulseGlow 1.7s infinite;
       color: #229a53; font-size:2.1em;
    }
    @keyframes pulseGlow {
      0% { text-shadow: 0 0 0px #22aa5880;}
      70% { text-shadow: 0 0 16px #22aa58b8;}
      100% { text-shadow: 0 0 0px #22aa5880;}
    }
    </style>
""", unsafe_allow_html=True)

# --- HEADER & LOGOS ---
left, center, right = st.columns([2, 6, 2])
with left:
    st.image("logo1.png", width=140)
with right:
    st.image("logo2.png", width=140)
with center:
    st.markdown("""
    <h1 style='text-align:center; margin-bottom:0.1em;'>HMV Fair Quote Validation Tool</h1>
    <hr style='border:2px solid #3498db; border-radius:5px; margin-bottom:1.2em;'>
    """, unsafe_allow_html=True)

# --- DATA PATH ---
DATA_PATH = 'hmv_data.xlsx'

# --- Load historic data from the single sheet
def load_historic_data(path):
    if not os.path.exists(path):
        st.error(f"Historical data file '{path}' not found.")
        st.stop()
    try:
        df = pd.read_excel(path)
        return df
    except Exception as e:
        st.error(f"Error reading data: {e}")
        st.stop()

df = load_historic_data(DATA_PATH)

# --- Data normalization ---
def normalize_text(text):
    if pd.isna(text): return ""
    text = str(text).upper()
    text = re.sub(r'\b\d{1,2}[-/]\d{1,2}[-/]\d{2,4}\b', '', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text

df['Normalized Corrective Action'] = df['Corrective Action'].apply(normalize_text)
df['Normalized Discrepancy'] = df['Description'].apply(lambda x: normalize_text(str(x).replace("(FOR REFERENCE ONLY)", "")))
df['Combined Key'] = df['Normalized Discrepancy'] + " | " + df['Normalized Corrective Action']

# --- Clustering similar keys (fuzzy matching) ---
clusters = {}
for key in df['Combined Key'].unique():
    if not key:
        continue
    for rep in clusters:
        if fuzz.token_set_ratio(key, rep) >= 90:
            clusters[rep].append(key)
            break
    else:
        clusters[key] = [key]

key_to_rep = {k: r for r, lst in clusters.items() for k in lst}
df['Cluster Key'] = df['Combined Key'].map(key_to_rep)

# --- Calculate average historic hours per cluster ---
hours = df.groupby('Cluster Key')['Total Hours'].agg(['mean', 'count']).reset_index()
hours.columns = ['Cluster Key', 'Actual Historic Hours', 'Occurrences']
df = df.merge(hours, on='Cluster Key', how='left')
df['Fair Quote (hrs)'] = df['Actual Historic Hours'].round(2)

# --- User Input Form ---
st.markdown("""
<div class="form-container" style="background:#eaf3fa;padding:1.6em 2.4em; border-radius:14px; box-shadow:0 2px 12px rgba(52,152,219,0.12);">
<h4 style='color:#1e4669; margin-bottom:12px;'>Enter Maintenance Quote Details</h4>
""", unsafe_allow_html=True)

with st.form("quote_form", clear_on_submit=False):
    col1, col2 = st.columns(2)
    with col1:
        discrepancy_input = st.text_area("Description of Non-Routine", height=90,
                                         placeholder="Describe the issue or discrepancy...")
    with col2:
        corrective_input = st.text_area("Corrective Action", height=90,
                                        placeholder="Describe the corrective action taken...")
    supplier_hours = st.number_input("Supplier Quoted Hours", min_value=0.0, step=0.1)
    submit = st.form_submit_button("🔍 Analyze Quote", use_container_width=True)
st.markdown("</div>", unsafe_allow_html=True)

# --- Helper functions ---
def semantic_overlap(a, b):
    if not a or not b:
        return 0
    matcher = difflib.SequenceMatcher(None, a.split(), b.split())
    return matcher.ratio() * 100

def get_decision_conclusion(supplier, fair):
    if fair == 0 or pd.isna(fair):
        percent_diff = "N/A"
        diff_class = "diff-neutral"
    else:
        percent_diff = ((supplier - fair) / fair) * 100
        if percent_diff < 0:
            diff_class = "diff-negative"
        elif abs(percent_diff) <= 5:
            diff_class = "diff-neutral"
        else:
            diff_class = "diff-positive"
    if fair == 0 or pd.isna(fair):
        percent_display = "N/A (no historical data)"
    else:
        sign = "+" if percent_diff >= 0 else ""
        percent_display = f"{sign}{percent_diff:.1f}%"
    if fair == 0 or pd.isna(fair):
        return ("No historical data available — manual review recommended.", "fail", percent_display, diff_class)
    if supplier < fair:
        return ("FAIR QUOTE: Supplier is below historic average. Consider approving.", "success", percent_display, diff_class)
    elif abs(supplier - fair) / fair <= 0.05:
        return ("IN EXPECTED RANGE (±5%). Consider approving.", "warn", percent_display, diff_class)
    else:
        return ("HIGHER THAN HISTORIC — Needs BP review.", "fail", percent_display, diff_class)

def highlight_diff(text, ref):
    ref_words = set(ref.split())
    return " ".join([f"<b><span style='color:#e67e22'>{w}</span></b>" if w not in ref_words else w for w in text.split()])


# --- Append new entry to Excel (same sheet) ---
def append_new_entry_to_excel(path, new_row):
    # Read existing workbook with openpyxl
    wb = openpyxl.load_workbook(path)
    ws = wb.active  # using single sheet

    # Find last row with data
    last_row = ws.max_row + 1

    # Columns header from first row
    headers = [cell.value for cell in ws[1]]

    # Align new_row to headers (fill missing columns with empty string)
    row_values = [new_row.get(col, "") for col in headers]

    # Append the row_values to next empty row
    ws.append(row_values)
    
    # Save workbook
    wb.save(path)

# --- Main logic ---
if submit and discrepancy_input and corrective_input:
    with st.spinner("🔎 Checking historic invoices..."):
        norm_disc = normalize_text(discrepancy_input.replace("(FOR REFERENCE ONLY)", ""))
        norm_corr = normalize_text(corrective_input)
        combined_input = norm_disc + " | " + norm_corr

        exact = df[df['Combined Key'] == combined_input]

        def total_similarity(row):
            d_ov = semantic_overlap(norm_disc, row['Normalized Discrepancy'])
            c_ov = semantic_overlap(norm_corr, row['Normalized Corrective Action'])
            return (d_ov + c_ov) / 2

        df['Overlap'] = df.apply(total_similarity, axis=1)
        approx = df[(df['Overlap'] >= 55) & (df['Combined Key'] != combined_input)]
        top2 = approx.sort_values(by='Overlap', ascending=False).head(2)
        closest = df[df['Overlap'] < 55].sort_values(by='Overlap', ascending=False).head(1)
        time.sleep(0.5)

    st.markdown("<hr>", unsafe_allow_html=True)

    # Display conclusions and metrics
    if not exact.empty:
        row = exact.iloc[0]
        conclusion, style_class, percent_diff, diff_class = get_decision_conclusion(supplier_hours, row['Fair Quote (hrs)'])
        st.markdown(f"""
            <div class="conclusion-animated {style_class}">
                <div class="conclusion-head">{'✅' if style_class=='success' else '⚠️' if style_class=='warn' else '⛔️'} {conclusion}</div>
                <div style="font-size:1.08em;">
                    Supplier quoted <b>{supplier_hours:.2f} hrs</b><br>
                    Historic average: <b>{row['Fair Quote (hrs)']:.2f} hrs</b><br>
                    <span style="color:#579;">{percent_diff} difference</span> &nbsp;|&nbsp; <b>Exact match found.</b>
                </div>
            </div>
        """, unsafe_allow_html=True)
        c1, c2, c3 = st.columns(3)
        with c1:
            st.markdown(f"<div class='metric-card'><span class='metric-label'>Historical (Fair) Hours</span><div class='metric-value'>{row['Fair Quote (hrs)']:.2f}</div></div>",unsafe_allow_html=True)
        with c2:
            st.markdown(f"<div class='metric-card'><span class='metric-label'>Supplier Quoted Hours</span><div class='metric-value'>{supplier_hours:.2f}</div></div>",unsafe_allow_html=True)
        with c3:
            st.markdown(f"<div class='metric-card'><span class='metric-label'>% Difference</span><div class='metric-value {diff_class}'>{percent_diff}</div></div>",unsafe_allow_html=True)
        st.success("Exact historic match found")
        st.dataframe(exact[['Description', 'Corrective Action', 'Actual Historic Hours', 'Fair Quote (hrs)', 'Occurrences']])
    elif not top2.empty:
        row = top2.iloc[0]
        conclusion, style_class, percent_diff, diff_class = get_decision_conclusion(supplier_hours, row['Fair Quote (hrs)'])
        st.markdown(f"""
            <div class="conclusion-animated {style_class}">
                <div class="conclusion-head">{'🟡' if style_class=='warn' else '⛔️'} {conclusion}</div>
                <div style="font-size:1.08em;">
                    Supplier quoted <b>{supplier_hours:.2f} hrs</b><br>
                    Historic average: <b>{row['Fair Quote (hrs)']:.2f} hrs</b><br>
                    <span style="color:#579;">{percent_diff} difference</span> &nbsp;|&nbsp; <b>Approximate match (≥55%)</b>
                </div>
            </div>
        """, unsafe_allow_html=True)
        c1, c2, c3 = st.columns(3)
        with c1:
            st.markdown(f"<div class='metric-card'><span class='metric-label'>Historical (Fair) Hours</span><div class='metric-value'>{row['Fair Quote (hrs)']:.2f}</div></div>",unsafe_allow_html=True)
        with c2:
            st.markdown(f"<div class='metric-card'><span class='metric-label'>Supplier Quoted Hours</span><div class='metric-value'>{supplier_hours:.2f}</div></div>",unsafe_allow_html=True)
        with c3:
            st.markdown(f"<div class='metric-card'><span class='metric-label'>% Difference</span><div class='metric-value {diff_class}'>{percent_diff}</div></div>",unsafe_allow_html=True)
        st.info("Top approximate matches (≥55% similarity):")
        rows = []
        for _, trow in top2.iterrows():
            rows.append({
                'Description': highlight_diff(trow['Normalized Discrepancy'], norm_disc),
                'Corrective Action': highlight_diff(trow['Normalized Corrective Action'], norm_corr),
                'Historic Hours': f"{trow['Actual Historic Hours']:.2f}",
                'Fair Quote (hrs)': f"{trow['Fair Quote (hrs)']:.2f}",
                'Occurrences': trow['Occurrences'],
                'Overlap %': f"{trow['Overlap']:.1f}%"
            })
        html_table = """
            <table style="width:100%;" class="result-table">
            <tr><th>Description</th><th>Corrective</th><th>Historic Hours</th><th>Fair Quote (hrs)</th><th>Occurrences</th><th>Overlap %</th></tr>
        """
        for row in rows:
            html_table += f"<tr><td>{row['Description']}</td><td>{row['Corrective Action']}</td><td>{row['Historic Hours']}</td><td>{row['Fair Quote (hrs)']}</td><td>{row['Occurrences']}</td><td>{row['Overlap %']}</td></tr>"
        html_table += "</table>"
        st.markdown(html_table, unsafe_allow_html=True)
    elif not closest.empty:
        row = closest.iloc[0]
        conclusion, style_class, percent_diff, diff_class = get_decision_conclusion(supplier_hours, row['Fair Quote (hrs)'])
        st.markdown(f"""
            <div class="conclusion-animated fail">
                <div class="conclusion-head">⛔️ {conclusion}</div>
                <div style="font-size:1.05em;">
                    Supplier quoted <b>{supplier_hours:.2f} hrs</b><br>
                    Historic average: <b>{row['Fair Quote (hrs)']:.2f} hrs</b><br>
                    <span style="color:#579;">{percent_diff} difference</span> &nbsp;|&nbsp; <b>Only distant/low-similarity match shown</b>
                </div>
            </div>
        """, unsafe_allow_html=True)
        c1, c2, c3 = st.columns(3)
        with c1:
            st.markdown(f"<div class='metric-card'><span class='metric-label'>Historical (Fair) Hours</span><div class='metric-value'>{row['Fair Quote (hrs)']:.2f}</div></div>",unsafe_allow_html=True)
        with c2:
            st.markdown(f"<div class='metric-card'><span class='metric-label'>Supplier Quoted Hours</span><div class='metric-value'>{supplier_hours:.2f}</div></div>",unsafe_allow_html=True)
        with c3:
            st.markdown(f"<div class='metric-card'><span class='metric-label'>% Difference</span><div class='metric-value {diff_class}'>{percent_diff}</div></div>",unsafe_allow_html=True)
        st.warning("No close matches found. Only nearest reference below.", icon="⚠️")
        st.markdown(f"""
            <table class="result-table" style="width:100%;">
            <tr><th>Description</th><th>Corrective Action</th><th>Historic Hours</th><th>Fair Quote (hrs)</th><th>Occurrences</th><th>Overlap %</th></tr>
            <tr>
            <td>{row['Normalized Discrepancy']}</td>
            <td>{row['Normalized Corrective Action']}</td>
            <td>{row['Actual Historic Hours']:.2f}</td>
            <td>{row['Fair Quote (hrs)']:.2f}</td>
            <td>{row['Occurrences']}</td>
            <td>{row['Overlap']:.1f}%</td>
            </tr></table>
        """, unsafe_allow_html=True)

        st.info(
            "❓ <b>No reliable or similar past instance found for this combination.<br>"
            "If this is a valid quote, you can contribute it for future reference:</b>",
            icon="🗂"
        )
        if st.button("➕ Add this Quote as a New Historical Instance"):
            new_entry = {
                'Description': discrepancy_input,
                'Corrective Action': corrective_input,
                'Total Hours': supplier_hours,
                'Year': pd.to_datetime("today").year
            }
            try:
                append_new_entry_to_excel(DATA_PATH, new_entry)
                st.success("✅ Added new quote for future reference!")
                conflitto.confetti()
            except Exception as e:
                st.error(f"Failed to save new instance. Error: {e}")
    else:
        st.info("No data found. Please enter details.")

else:
    st.info("Enter a quote description, corrective action, and hours to begin.")
